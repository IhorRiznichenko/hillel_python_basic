import openpyxl
from person import Person


def write_id_person():
    name_column = ["Имя", "Отчество", "Фамилия", "Дата рождения", "Дата смерти", "Пол", "Возраст"]
    filename = input("Введите имя существующего файла или новое имя файла: ")
    while not filename:
        filename = input("Имя файла не может быть пустым. Пожалуйста, введите имя файла: ")
    try:
        wb = openpyxl.load_workbook(f"{filename}.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.create_sheet(title="First list", index=0)
        wb.remove(wb['Sheet'])
    sheet = wb["First list"]
    rows = sheet.max_row
    name = input("Введите имя: ").title()
    while not name:
        name = input("Имя  не может быть пустым. Пожалуйста, введите имя : ")
    middle_name = input("Введите отчество (необязательно): ").title()
    surname = input("Введите фамилию (необязательно): ").title()

    while True:
        data_of_birthdata = input("Введите дату рождения в формате: дд.мм.гггг; дд мм гггг; дд/мм/гггг; д-м-гггг : ")
        try:
            Person.create_date(data_of_birthdata)
            break
        except ValueError as a:
            print(a)
            continue
    while True:
        data_of_death = input("Введите дату смерти в формате: дд.мм.гггг; дд мм гггг; дд/мм/гггг; д-м-гггг (необязательно, оставьте поле пустым, если живет): ")
        try:
            Person.create_date_1(data_of_death)
            break
        except ValueError as a:
            print(a)
            continue
    while True:
        sex = input("Введите пол (м/ж): ")
        try:
            Person.validate_sex(sex)
            break
        except ValueError as a:
            print(a)
    person = Person(name, data_of_birthdata, sex, data_of_death, surname, middle_name)
    for col_index, value in enumerate(name_column):
        cell = sheet.cell(row=1, column=col_index + 1)
        cell.value = value
    for col_index, value in enumerate(person.create_list()):
        cell = sheet.cell(row=rows + 1, column=col_index + 1)
        cell.value = value
    wb.save(f"{filename}.xlsx")



def search_id_person():
    while True:
        filename = input("Введите имя файла: ")
        try:
            wb = openpyxl.load_workbook(f"{filename}.xlsx")
        except FileNotFoundError:
            print(f"Файл {filename}.xlsx не найден. Попробуйте еще раз.")
            continue
        input_item_search = input("Введите данные из файла: ").replace(" ", "").lower()
        if not input_item_search:
            print("Вы не ввели данные из файла. Попробуйте еще раз.")
            continue
        fields_rows = []
        sheet = wb["First list"]
        rows = sheet.max_row
        cols = sheet.max_column
        for i in range(1, rows+1):
            value_row = []
            for j in range(1, cols+1):
                cell = sheet.cell(row=i, column=j)
                value = str(cell.value)
                if i == 1:
                    continue
                else:
                    value_row.append(value)
            fields_rows.append(value_row)
        wb.save(f"{filename}.xlsx")
        count = 0
        for rows in fields_rows:
            try:
                if "".join(rows[0:3]).lower().index(input_item_search) == 0 or \
                        "".join(rows[0:3]).lower().index(input_item_search) == 1:
                    print(Person.print_list(rows))
                    count += 1
                    continue
            except ValueError:
                continue
        if count == 0:
            choice = input("Данные не найдено. Хотите добавить новую запись? (да/нет): ").lower()
            if choice == "да":
                write_id_person()
                break
            else:
                continue
        else:
            choice = input("Хотите добавить новую запись? (да/нет): ").lower()
            if choice == "да":
                write_id_person()
                break
            else:
                break

while True:
    function = {"1": write_id_person, "2": search_id_person}
    print('''
1: Создать файл с данными или Добавить данные в существующий файл, нажмите 1
2: Найти данные в существующем файле, нажмите 2
3: Если вы хотите выйти нажмите 3''')
    selection = input("Введите ваш выбор: ")
    if not selection.isdigit() or int(selection) == 0 or int(selection) > 3:
        print("Ви ввели некоректне значення, попробуйте еще")
        print("-" * 50)
        continue
    elif int(selection) == 3:
        break
    function[selection]()
    print("-" * 100 + "\n")
