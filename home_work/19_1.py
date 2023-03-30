import csv
from openpyxl import Workbook

with open('my_dict.csv', newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile)
    data = [row for row in reader]
for row in data:
    del row[2]
transposed_data = list(map(list, zip(*data)))

transposed_data.insert(0, [])

workbook = Workbook()
worksheet = workbook.active
for row in transposed_data:
    worksheet.append(row)
worksheet.cell(row=1, column=2, value='Person1')
worksheet.cell(row=1, column=3, value='Person2')
worksheet.cell(row=1, column=4, value='Person3')
worksheet.cell(row=1, column=5, value='Person4')
worksheet.cell(row=1, column=6, value='Person5')
workbook.save('my_dict.xlsx')
